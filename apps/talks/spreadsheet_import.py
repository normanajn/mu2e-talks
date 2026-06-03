import json
import re
from datetime import date, datetime
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path

from django.db import transaction

from apps.accounts.models import Institution, InstitutionAlias, User, UserAlias

from .models import Conference, Talk


class TalkSpreadsheetImportError(ValueError):
    pass


SPREADSHEET_NAME = 'TalksSpreadsheet-DocDB3701.xlsx'
HEADER_ALIASES = {
    'date': 'date',
    'speaker': 'speaker_last_name',
    'home institution': 'home_institution',
    'venue': 'venue',
    'title': 'title',
    'length (min)': 'length',
    'length min': 'length',
    'type': 'type',
    'doc-db number': 'docdb',
    'doc db number': 'docdb',
    'doc-db (pwd)': 'docdb_pwd',
    'doc db pwd': 'docdb_pwd',
    'doc-db(cert)': 'docdb_cert',
    'doc db cert': 'docdb_cert',
    'doc-db': 'docdb',
    'doc db': 'docdb',
    'conference web page': 'conference_url',
    'conference url': 'conference_url',
    'approved by committee?': 'committee_approved',
    'approved by committee': 'committee_approved',
    'mu2e-ii?': 'mu2e_program',
    'mu2e ii': 'mu2e_program',
    'mu2e-ii': 'mu2e_program',
    'mu2e or mu2e-ii': 'mu2e_program',
    'mu2e or mu2e ii': 'mu2e_program',
    'proceedings': 'proceedings_url',
    'arxiv': 'arxiv_url',
}
INSTITUTION_ALIASES = {
    'anl': 'Argonne National Laboratory',
    'argonne': 'Argonne National Laboratory',
    'bu': 'Boston University',
    'bnl': 'Brookhaven National Laboratory',
    'caltech': 'California Institute of Technology',
    'cal tech': 'California Institute of Technology',
    'fnal': 'Fermi National Accelerator Laboratory',
    'fermilab': 'Fermi National Accelerator Laboratory',
    'frascati': 'Laboratori Nazionali di Frascati',
    'hzdr': 'Helmholtz-Zentrum Dresden-Rossendorf',
    'northwestern': 'Northwestern University',
    'lbl': 'Lawrence Berkeley National Laboratory',
    'lbnl': 'Lawrence Berkeley National Laboratory',
    'lnf': 'Laboratori Nazionali di Frascati',
    'pisa': 'Universita di Pisa / Istituto Nazionale di Fisica Nucleare, Pisa',
    'ucberkeley': 'University of California, Berkeley',
    'uva': 'University of Virginia',
    'uv a': 'University of Virginia',
    'uv a.': 'University of Virginia',
    'umn': 'University of Minnesota',
    'uminn': 'University of Minnesota',
    'uminnesota': 'University of Minnesota',
}


def _normalize(value):
    value = str(value or '').strip().lower()
    value = value.replace('&', ' and ')
    return re.sub(r'[^a-z0-9]+', ' ', value).strip()


def _text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _as_int(value):
    text = _text(value)
    if not text:
        return None
    match = re.search(r'\d+', text)
    return int(match.group()) if match else None


def _yes(value):
    return _normalize(value) in {'1', 'true', 'yes', 'y', 'on', 'as delivered', 'as presented'}


def _bool_value(value):
    if isinstance(value, bool):
        return value
    return _yes(value)


def _docdb(value):
    text = _text(value)
    if not text:
        return ''
    match = re.search(r'\d+', text)
    return match.group() if match else text


def _type_from_raw(value):
    normalized = _normalize(value)
    if 'seminar' in normalized:
        return Talk.Type.SEMINAR
    if 'colloquium' in normalized:
        return Talk.Type.COLLOQUIUM
    if 'conference' in normalized:
        return Talk.Type.CONFERENCE
    if any(term in normalized for term in ('parallel', 'plenary', 'poster')):
        return Talk.Type.CONFERENCE
    return Talk.Type.OTHER


def _column_map(ws):
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), max_col=20, values_only=False):
        values = [_normalize(cell.value) for cell in row]
        if 'date' not in values or 'title' not in values:
            continue
        mapping = {}
        previous_key = None
        for col_index, cell in enumerate(row, start=1):
            header = _normalize(cell.value)
            key = HEADER_ALIASES.get(header)
            if key:
                mapping[key] = col_index
                previous_key = key
                continue
            # The first-name column is blank under the merged "Speaker" header.
            if previous_key == 'speaker_last_name' and not header:
                mapping['speaker_first_name'] = col_index
                previous_key = 'speaker_first_name'
        return row[0].row, mapping
    raise TalkSpreadsheetImportError(f'Sheet "{ws.title}" does not contain a recognized talk header row.')


def workbook_to_records(upload):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(upload, data_only=True, read_only=True)
    except Exception as exc:
        raise TalkSpreadsheetImportError(f'Could not read spreadsheet: {exc}') from exc

    records = []
    for ws in wb.worksheets:
        header_row, columns = _column_map(ws)
        for source_row, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_col=20, values_only=True),
            start=header_row + 1,
        ):

            def value(key):
                col = columns.get(key)
                return row[col - 1] if col else None

            talk_date = _as_date(value('date'))
            title = _text(value('title'))
            venue = _text(value('venue'))
            last = _text(value('speaker_last_name')).rstrip('?').strip()
            first = _text(value('speaker_first_name')).rstrip('?').strip()
            if ws.title == '2013-2015':
                first, last = last, first
            if not (talk_date or title or venue or last or first):
                continue
            docdb_pwd = _docdb(value('docdb_pwd') or value('docdb'))
            docdb_cert = _docdb(value('docdb_cert'))
            records.append({
            'source_spreadsheet': Path(getattr(upload, 'name', '') or SPREADSHEET_NAME).name,
                'source_sheet': ws.title,
                'source_row': source_row,
                'presentation_date': talk_date.isoformat() if talk_date else '',
                'speaker_last_name': last,
                'speaker_first_name': first,
                'speaker_home_institution_raw': _text(value('home_institution')),
                'conference_title': venue,
                'conference_url': _text(value('conference_url')),
                'talk_title': title or venue or 'Untitled talk',
                'duration_raw': _text(value('length')),
                'duration_minutes': _as_int(value('length')),
                'spreadsheet_type_raw': _text(value('type')),
                'type': _type_from_raw(value('type')),
            'plenary': 'plenary' in _normalize(value('type')),
            'parallel': 'parallel' in _normalize(value('type')),
                'docdb_number': docdb_cert or docdb_pwd,
                'docdb_password_number': docdb_pwd,
                'docdb_certificate_number': docdb_cert,
                'committee_approved_raw': _text(value('committee_approved')),
                'final_approval': _yes(value('committee_approved')),
                'mu2e_program': _text(value('mu2e_program')),
                'proceedings_url': _text(value('proceedings_url')),
                'arxiv_url': _text(value('arxiv_url')),
            })
    return records


def records_to_json(records):
    return json.dumps(records, ensure_ascii=False, indent=2)


def records_from_json(upload):
    try:
        data = json.load(upload)
    except Exception as exc:
        raise TalkSpreadsheetImportError(f'Could not read JSON: {exc}') from exc
    if not isinstance(data, list):
        raise TalkSpreadsheetImportError('Talk import JSON must contain a list of records.')
    return data


def _match_institution(raw):
    normalized = _normalize(raw)
    if not normalized:
        return None, 'no institution in spreadsheet'

    alias_name = INSTITUTION_ALIASES.get(normalized)
    candidates = Institution.objects.all()
    if alias_name:
        exact = candidates.filter(name__iexact=alias_name).first()
        if exact:
            return exact, ''

    exact = candidates.filter(name__iexact=raw).first()
    if exact:
        return exact, ''
    for field in ('collaboration_code', 'collaboration_number'):
        match = candidates.filter(**{f'{field}__iexact': raw}).first()
        if match:
            return match, ''

    alias = (
        InstitutionAlias.objects.select_related('institution')
        .filter(normalized_alias=normalized, is_active=True)
        .first()
    )
    if alias:
        return alias.institution, ''

    best = None
    best_score = 0.0
    for institution in candidates:
        score = SequenceMatcher(None, normalized, _normalize(institution.name)).ratio()
        if score > best_score:
            best = institution
            best_score = score
    if best and best_score >= 0.82:
        return best, f'institution fuzzy matched at {best_score:.2f}'
    return None, f'institution not matched: {raw}'


@lru_cache(maxsize=1)
def _user_match_candidates():
    candidates = []
    for user in User.objects.select_related('institution'):
        values = tuple(
            value
            for value in (
                _normalize(user.display_name),
                _normalize(user.get_full_name()),
                _normalize(user.username),
                _normalize(user.email.split('@')[0] if user.email else ''),
                _normalize(user.contact_email.split('@')[0] if user.contact_email else ''),
            )
            if value
        )
        if values:
            candidates.append((user.pk, user.institution_id, values))
    return tuple(candidates)


def _score_user_alias(expected, expected_reverse, values):
    expected_tokens = set(expected.split())
    reverse_tokens = set(expected_reverse.split())
    best = 0.0
    for value in values:
        value_tokens = set(value.split())
        if expected == value or expected_reverse == value:
            return 1.0
        if not (expected_tokens & value_tokens or reverse_tokens & value_tokens):
            continue
        best = max(
            best,
            SequenceMatcher(None, expected, value).ratio(),
            SequenceMatcher(None, expected_reverse, value).ratio(),
        )
    return best


def _match_user(first, last, institution):
    first_norm = _normalize(first)
    last_norm = _normalize(last)
    if not (first_norm or last_norm):
        return None, 'no speaker name in spreadsheet'
    expected = f'{first_norm} {last_norm}'.strip()
    expected_reverse = f'{last_norm} {first_norm}'.strip()

    alias_qs = UserAlias.objects.select_related('user').filter(
        normalized_alias__in={expected, expected_reverse},
        is_active=True,
    )
    if institution:
        institution_alias = alias_qs.filter(institution=institution).first()
        if institution_alias:
            return institution_alias.user, ''
    alias = alias_qs.filter(institution__isnull=True).first() or alias_qs.first()
    if alias:
        return alias.user, ''

    if len(expected) > 128 or len(expected_reverse) > 128:
        return None, f'user not matched: {first} {last}'.strip()

    best_user_id = None
    best_score = 0.0
    for user_id, institution_id, values in _user_match_candidates():
        score = _score_user_alias(expected, expected_reverse, values)
        if institution_id and institution and institution_id == institution.id:
            score += 0.05
        if score > best_score:
            best_user_id = user_id
            best_score = score
    if best_user_id and best_score >= 0.86:
        note = '' if best_score >= 0.95 else f'user fuzzy matched at {best_score:.2f}'
        return User.objects.get(pk=best_user_id), note
    return None, f'user not matched: {first} {last}'.strip()


def _conference_for(record):
    title = record.get('conference_title', '').strip()
    if not title:
        title = 'Unknown conference'
    conference, _ = Conference.objects.get_or_create(
        title=title,
        defaults={
            'start_date': _as_date(record.get('presentation_date')),
            'end_date': _as_date(record.get('presentation_date')),
            'url': record.get('conference_url', '').strip(),
        },
    )
    changed = []
    if record.get('conference_url') and not conference.url:
        conference.url = record['conference_url'].strip()
        changed.append('url')
    if record.get('presentation_date') and not conference.start_date:
        conference.start_date = _as_date(record['presentation_date'])
        conference.end_date = _as_date(record['presentation_date'])
        changed += ['start_date', 'end_date']
    if changed:
        conference.save(update_fields=changed + ['updated_at'])
    return conference


@transaction.atomic
def import_talk_records(records, created_by):
    _user_match_candidates.cache_clear()
    counts = {'created': 0, 'updated': 0, 'matched_users': 0, 'unmatched_users': 0,
              'matched_institutions': 0, 'unmatched_institutions': 0}
    for record in records:
        institution, institution_note = _match_institution(record.get('speaker_home_institution_raw', ''))
        user, user_note = _match_user(
            record.get('speaker_first_name', ''),
            record.get('speaker_last_name', ''),
            institution,
        )
        counts['matched_institutions' if institution else 'unmatched_institutions'] += 1
        counts['matched_users' if user else 'unmatched_users'] += 1
        conference = _conference_for(record)
        notes = '; '.join(note for note in (institution_note, user_note) if note)
        lookup = {
            'source_spreadsheet': record.get('source_spreadsheet', SPREADSHEET_NAME),
            'source_sheet': record.get('source_sheet', ''),
            'source_row': int(record.get('source_row') or 0),
        }
        defaults = {
            'conference': conference,
            'talk_title': record.get('talk_title', '').strip() or 'Untitled talk',
            'presentation_date': _as_date(record.get('presentation_date')),
            'type': record.get('type') or Talk.Type.OTHER,
            'spreadsheet_type_raw': record.get('spreadsheet_type_raw', '').strip(),
            'docdb_number': record.get('docdb_number', '').strip(),
            'docdb_password_number': record.get('docdb_password_number', '').strip(),
            'docdb_certificate_number': record.get('docdb_certificate_number', '').strip(),
            'plenary': _bool_value(record.get('plenary')),
            'parallel': _bool_value(record.get('parallel')),
            'assigned_to': user,
            'speaker_first_name': record.get('speaker_first_name', '').strip(),
            'speaker_last_name': record.get('speaker_last_name', '').strip(),
            'speaker_home_institution_raw': record.get('speaker_home_institution_raw', '').strip(),
            'speaker_institution': institution,
            'duration_minutes': record.get('duration_minutes') or None,
            'duration_raw': record.get('duration_raw', '').strip(),
            'final_approval': _bool_value(record.get('final_approval')),
            'committee_approved_raw': record.get('committee_approved_raw', '').strip(),
            'complete_given': True,
            'mu2e_program': record.get('mu2e_program', '').strip(),
            'proceedings_url': record.get('proceedings_url', '').strip(),
            'arxiv_url': record.get('arxiv_url', '').strip(),
            'status': Talk.Status.ACTIVE,
            'created_by': created_by,
            'spreadsheet_import_notes': notes,
        }
        talk, created = Talk.objects.update_or_create(defaults=defaults, **lookup)
        counts['created' if created else 'updated'] += 1
    return counts


def workbook_path_to_json(input_path, output_path):
    with Path(input_path).open('rb') as handle:
        records = workbook_to_records(handle)
    Path(output_path).write_text(records_to_json(records), encoding='utf-8')
    return len(records)
