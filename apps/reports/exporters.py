import csv
import io
import json

from django.http import HttpResponse
from django.utils import timezone

_registry: dict = {}
SPREADSHEET_FORMULA_PREFIXES = ('=', '+', '-', '@')


def register(fmt: str):
    def decorator(fn):
        _registry[fmt] = fn
        return fn
    return decorator


def get(fmt: str):
    return _registry.get(fmt)


def available() -> list[str]:
    return list(_registry.keys())


def _filename(fmt: str) -> str:
    ts = timezone.now().strftime('%Y%m%d_%H%M')
    return f'mu2e_talks_report_{ts}.{fmt}'


HEADERS = [
    'id',
    'status',
    'conference_title',
    'talk_title',
    'presentation_date',
    'type',
    'spreadsheet_type_raw',
    'conference_start',
    'conference_end',
    'conference_url',
    'docdb_number',
    'docdb_password_number',
    'docdb_certificate_number',
    'plenary',
    'parallel',
    'assigned_to',
    'speaker_first_name',
    'speaker_last_name',
    'speaker_home_institution',
    'speaker_home_institution_raw',
    'duration_minutes',
    'duration_raw',
    'practice_talk_date',
    'practice_talk_complete',
    'final_approval',
    'committee_approved_raw',
    'complete_given',
    'mu2e_program',
    'proceedings_url',
    'arxiv_url',
    'comments',
]


def _rows(qs):
    for talk in qs.select_related('conference', 'assigned_to', 'speaker_institution'):
        conference = talk.conference
        yield {
            'id': talk.pk,
            'status': talk.get_status_display(),
            'conference_title': conference.title if conference else '',
            'talk_title': talk.talk_title,
            'presentation_date': talk.presentation_date.isoformat() if talk.presentation_date else '',
            'type': talk.get_type_display(),
            'spreadsheet_type_raw': talk.spreadsheet_type_raw,
            'conference_start': conference.start_date.isoformat() if conference and conference.start_date else '',
            'conference_end': conference.end_date.isoformat() if conference and conference.end_date else '',
            'conference_url': conference.url if conference else '',
            'docdb_number': talk.docdb_number,
            'docdb_password_number': talk.docdb_password_number,
            'docdb_certificate_number': talk.docdb_certificate_number,
            'plenary': 'yes' if talk.plenary else 'no',
            'parallel': 'yes' if talk.parallel else 'no',
            'assigned_to': str(talk.assigned_to) if talk.assigned_to else '',
            'speaker_first_name': talk.speaker_first_name,
            'speaker_last_name': talk.speaker_last_name,
            'speaker_home_institution': str(talk.speaker_institution) if talk.speaker_institution else '',
            'speaker_home_institution_raw': talk.speaker_home_institution_raw,
            'duration_minutes': talk.duration_minutes or '',
            'duration_raw': talk.duration_raw,
            'practice_talk_date': talk.practice_talk_date.isoformat() if talk.practice_talk_date else '',
            'practice_talk_complete': 'yes' if talk.practice_talk_complete else 'no',
            'final_approval': 'yes' if talk.final_approval else 'no',
            'committee_approved_raw': talk.committee_approved_raw,
            'complete_given': 'yes' if talk.complete_given else 'no',
            'mu2e_program': talk.mu2e_program,
            'proceedings_url': talk.proceedings_url,
            'arxiv_url': talk.arxiv_url,
            'comments': talk.comments,
        }


def _spreadsheet_safe(value):
    if isinstance(value, str) and value.startswith(SPREADSHEET_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _spreadsheet_rows(qs):
    for row in _rows(qs):
        yield {key: _spreadsheet_safe(value) for key, value in row.items()}


@register('txt')
def export_txt(qs) -> HttpResponse:
    lines = []
    for row in _rows(qs):
        lines.append(
            f"[{row['status']}] {row['talk_title']}\n"
            f"  Type: {row['type']}\n"
            f"  Conference: {row['conference_title']} ({row['conference_start']} - {row['conference_end']})\n"
            f"  Assigned: {row['assigned_to'] or 'Unassigned'} | DocDB: {row['docdb_number'] or '-'} | "
            f"Plenary={row['plenary']} | Parallel={row['parallel']}\n"
            f"  Practice: {row['practice_talk_date'] or '-'} complete={row['practice_talk_complete']} | "
            f"Final approval={row['final_approval']} | Given={row['complete_given']}\n"
            f"  {row['comments']}\n"
        )
    body = '\n'.join(lines) or '(no talks matched)\n'
    resp = HttpResponse(body, content_type='text/plain; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("txt")}"'
    return resp


@register('md')
def export_md(qs) -> HttpResponse:
    sections = []
    for row in _rows(qs):
        sections.append(
            f"## {row['talk_title']}\n\n"
            f"**Status:** {row['status']}  \n"
            f"**Type:** {row['type']}  \n"
            f"**Conference:** {row['conference_title']}  \n"
            f"**Conference dates:** {row['conference_start']} - {row['conference_end']}  \n"
            f"**Assigned to:** {row['assigned_to'] or 'Unassigned'}  \n"
            f"**DocDB Number:** {row['docdb_number'] or '-'}  \n"
            f"**Plenary:** {row['plenary']}  \n"
            f"**Parallel:** {row['parallel']}  \n"
            f"**Practice talk:** {row['practice_talk_date'] or '-'}  \n"
            f"**Practice complete:** {row['practice_talk_complete']}  \n"
            f"**Final approval:** {row['final_approval']}  \n"
            f"**Complete/Given:** {row['complete_given']}\n\n"
            f"{row['comments']}"
        )
    body = '\n\n---\n\n'.join(sections) or '_(no talks matched)_'
    resp = HttpResponse(body, content_type='text/markdown; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("md")}"'
    return resp


@register('csv')
def export_csv(qs) -> HttpResponse:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=HEADERS)
    writer.writeheader()
    for row in _spreadsheet_rows(qs):
        writer.writerow(row)
    resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("csv")}"'
    return resp


@register('json')
def export_json(qs) -> HttpResponse:
    body = json.dumps(list(_rows(qs)), ensure_ascii=False, indent=2)
    resp = HttpResponse(body, content_type='application/json; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("json")}"'
    return resp


@register('xlsx')
def export_xlsx(qs) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mu2eTalks Report'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F766E')
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, row in enumerate(_spreadsheet_rows(qs), start=2):
        for col_idx, key in enumerate(HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{_filename("xlsx")}"'
    return resp


@register('pdf')
def export_pdf(qs) -> HttpResponse:
    from ._pdf import talks_to_pdf

    rows = list(_rows(qs))
    ts = timezone.now().strftime('%Y-%m-%d %H:%M')
    pdf_bytes = talks_to_pdf(
        rows=rows,
        title='Mu2eTalks Report',
        meta=f'Generated {ts} UTC  -  {len(rows)} talks',
    )
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{_filename("pdf")}"'
    return resp
